import openpyxl
workbook=openpyxl.load_workbook("book2.xlsx")
sheet=workbook.sheetnames
# print(sheet)
# print(workbook.active.title)
# print(workbook.sheetnames)

# sheet1=workbook['Student']
# value1=sheet1['A3'].value
# print(value1)

## method 1:
# print(workbook['Student']['A3'].value)

## method2:
# sheet1=workbook['Info']
# print(sheet1.cell(4,3).value)     #cell(row,column)

# #method 3:
# sheet2=workbook['product']
# print(sheet2.cell(row=4,column=3).value)

##Number of row & column:
sheet1=workbook['Info']

# rows=sheet1.max_row
# columns=sheet1.max_column
# print(rows)
# print(columns)

# #or

# rows=sheet1.max_row
# columns=sheet1.max_column
# print(f'The number of rows:',rows)           #f formated string
# print(f'The number of columns :',columns)

# #Read the value of all row and coulumns:
# rows=sheet1.max_row
# columns=sheet1.max_column
# for a in range(1,rows+1):
#     for b in range(1,columns+1):
#             print(sheet1.cell(a,b).value)

# sheet1=workbook['Info']
# sheet1.cell(row=13,column=1,value='Mrinank')
# sheet1.cell(row=13,column=2,value='Kolkata')
# sheet1.cell(row=13,column=3,value=21)
# workbook.save('Book2.xlsx')
# print('data Inserted succefully..!')





