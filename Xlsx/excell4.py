# function to remove empty rows
import openpyxl

def remove(sheet, row):

	for cell in row:

		if cell.value != None:
			return
	# get the row number from the first cell
	# and remove the row
	sheet.delete_rows(row[0].row, 1)


if __name__ == '__main__':

	path = 'C:\\Users\\AJAY KUMAR TIWARI\\OneDrive\\Desktop\\Python\\Xlsx\\delete_every_rows.xlsx'

	book = openpyxl.load_workbook(path)

	sheet = book['sheet1']

	print("Maximum rows before removing:", sheet.max_row)

	for row in sheet:
		remove(sheet,row)
		
	print("Maximum rows after removing:",sheet.max_row)
	
	# save the file to the path
	path = 'C:\\Users\\AJAY KUMAR TIWARI\\OneDrive\\Desktop\\Python\\Xlsx\\openpy11.xlsx'
	book.save(path)
