import openpyxl
workbook=openpyxl.load_workbook("book2.xlsx")


# sheet1=workbook['Info']
# sheet1.cell(row=6,column=1,value='Sujata')
# sheet1.cell(row=6,column=2,value='Delhi')
# sheet1.cell(row=6,column=3,value=15)
# workbook.save('book2.xlsx')
# print('data Inserted succefully..!')

#To save older datials with new one in a new File:
# sheet1=workbook['Info']
# sheet1.cell(row=14,column=1,value='Josh')  #if file is not present then it created by it slef.
# sheet1.cell(row=14,column=2,value='USA')
# sheet1.cell(row=14,column=3,value=48)
# workbook.save('Updated11.xlsx')
# print('data Udated succefully..!')


