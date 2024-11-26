import openpyxl as xl;


filename ="Book2.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[1]


filename1 ="new10.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active


mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		
		c = ws1.cell(row = i, column = j)

		# writing the read value to destination excel file
		ws2.cell(row = i, column = j).value = c.value

wb2.save(str(filename1))
