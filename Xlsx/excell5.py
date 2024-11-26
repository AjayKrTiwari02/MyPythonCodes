# continuously delete row 2 until there
# is only a single row left over

import openpyxl


def delete(sheet):



	while(sheet.max_row > 1):
		# this method removes the row 2
		sheet.delete_rows(2)

	return


if __name__ == '__main__':


	path = 'C:\\Users\\AJAY KUMAR TIWARI\\OneDrive\\Desktop\\Python\\Xlsx\\delete_every_rows.xlsx'

	book = openpyxl.load_workbook(path)

	# select the sheet
	sheet = book['sheet1']

	print("Maximum rows before removing:", sheet.max_row)

	delete(sheet)

	print("Maximum rows after removing:", sheet.max_row)

	# save the file to the path
	path = 'C:\\Users\\AJAY KUMAR TIWARI\\OneDrive\\Desktop\\Python\\Xlsx\\ABCDxyz.xlsx'
	book.save(path)
